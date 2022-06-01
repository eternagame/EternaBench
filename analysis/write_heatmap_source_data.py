import pandas as pd
import eternabench as eb
from openpyxl import load_workbook


def write_Fig_1A_1C_ED2():
    data = eb.load_CM_example_calculations()
    packages = eb.plot.get_packages() # dataframe containing information (full Title, color category) about each package

    example_package_list=['nupack_99','rnastructure', 'vienna_2', 'rnasoft_blstar','contrafold_2']
    titles = ['NUPACK 3.2.2: R = 0.699', 'RNAstructure: 0.730', 'ViennaRNA 2.4.14: 0.736', 'RNAsoft BLstar: 0.745', 'CONTRAfold 2: 0.750']

    # Reac data
    reac_dat = eb.plot.reactivity_heatmap_SOURCE_DATA(data, ind_range=[0,60])
    tmp = pd.DataFrame(reac_dat,columns=['reac_seqpos_%d'%i for i in range(79)] )

    with pd.ExcelWriter('SOURCE_DATA.xlsx',engine="openpyxl", mode='w') as writer:  
        tmp.to_excel(writer, sheet_name='Fig_1A')
        tmp.to_excel(writer, sheet_name='Fig_ED2')
        writer.save()    
        
    # Fig 1C example heatmaps
    for i, package in enumerate(example_package_list):
        dat = eb.plot.punpaired_heatmap_SOURCE_DATA(data, ind_range=[0,60], package=package)
        tmp = pd.DataFrame(dat,columns=['%s_seqpos_%d'%(package,i) for i in range(79)] )
        
        with pd.ExcelWriter('SOURCE_DATA.xlsx',engine="openpyxl", mode='a',if_sheet_exists='overlay' ) as writer: 
            if 'Fig_1C' in writer.book:
                maxrow=writer.book['Fig_1C'].max_row+1
            else:
                maxrow=0
            print(maxrow)
            tmp.to_excel(writer, sheet_name='Fig_1C', startrow=maxrow)

    # Fig ED2 all example heatmaps

    for j, pkg_kind in enumerate(packages.category.unique()):
        tmp = packages.loc[packages.category==pkg_kind]
        for package, row in tmp.iterrows():
            if 'p_'+package in data.keys():
                dat = eb.plot.punpaired_heatmap_SOURCE_DATA(data, ind_range=[0,60], package=package)
                tmp = pd.DataFrame(dat,columns=['%s_seqpos_%d'%(package,i) for i in range(79)] )
                with pd.ExcelWriter('SOURCE_DATA.xlsx',engine="openpyxl", mode='a',if_sheet_exists='overlay' ) as writer: 
                    maxrow=writer.book['Fig_ED2'].max_row+1
                    print(maxrow)
                    tmp.to_excel(writer, sheet_name='Fig_ED2', startrow=maxrow)

def write_Fig_ED4():
    example_package_list=['nupack_99','rnastructure', 'vienna_2', 'rnasoft_blstar','contrafold_2']

    df = pd.read_json('../data/EternaBench_ChemMapping_Filtered_10Jul2021.json.zip')
    df = df.loc[df.Dataset.str.startswith('Round')]
    arr = eb.plot.reactivity_heatmap_SOURCE_DATA(df, aspect='auto')
    tmp = pd.DataFrame(df,columns=['reac_seqpos_%d'%i for i in range(103)] )

    with pd.ExcelWriter('SOURCE_DATA_ED4.xlsx',engine="openpyxl", mode='w') as writer:  
        tmp.to_excel(writer, sheet_name='ED4')
        writer.save()    
        
    for i, package in enumerate(example_package_list):
        df = pd.read_json('../data/ChemMappingCalculations/CM_%s.json.zip'% package)
        df = df.loc[df.Dataset.str.startswith('Round')]
        arr = eb.plot.punpaired_heatmap_SOURCE_DATA(df, package=package,aspect='auto')
        
        tmp = pd.DataFrame(arr,columns=['%s_seqpos_%d'%(package,i) for i in range(103)] )

        with pd.ExcelWriter('SOURCE_DATA_ED4.xlsx',engine="openpyxl", mode='a',if_sheet_exists='overlay' ) as writer: 
                            maxrow=writer.book['ED4'].max_row+1
                            print(maxrow)
                            tmp.to_excel(writer, sheet_name='ED4', startrow=maxrow)

        
if __name__=='__main__':
    write_Fig_1A_1C_ED2()
    #write_Fig_ED4() didn't use this, uploaded raw .json.zip files instead
